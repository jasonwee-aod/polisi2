"""Extract text or build Anthropic content blocks from uploaded file attachments.

Supported file types:
- PDF: pypdf text extraction if available, otherwise Claude-native document block.
- DOCX: python-docx text extraction if available, otherwise plain note.
- XLSX: openpyxl to markdown table if available, otherwise plain note.
- CSV: stdlib csv to markdown table.
- TXT / MD: direct text inclusion.
- Images (PNG, JPG, GIF, WEBP): Anthropic image content block (Claude vision).
"""

from __future__ import annotations

import base64
import csv
import io
import logging
from dataclasses import dataclass

from polisi_api.models import FileAttachment

logger = logging.getLogger("polisi.file_processor")

# Content-type families -------------------------------------------------------

_IMAGE_TYPES = {
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/webp",
}

_PDF_TYPE = "application/pdf"

_DOCX_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

_XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

_CSV_TYPES = {
    "text/csv",
    "application/csv",
}

_TEXT_TYPES = {
    "text/plain",
    "text/markdown",
}

# Per-attachment token cost estimate added to the rate-limit budget.
TOKENS_PER_ATTACHMENT = 2000


@dataclass
class ProcessedAttachment:
    """Result of processing a single uploaded file."""

    filename: str
    text_content: str | None  # Extracted text (for text-based files)
    content_block: dict | None  # Anthropic API content block (for images / PDFs)


def process_attachments(attachments: list[FileAttachment]) -> list[ProcessedAttachment]:
    """Process uploaded files into text or Anthropic content blocks."""
    results: list[ProcessedAttachment] = []
    for att in attachments:
        try:
            results.append(_process_one(att))
        except Exception:
            logger.exception("Failed to process attachment %s", att.filename)
            results.append(
                ProcessedAttachment(
                    filename=att.filename,
                    text_content=f"[Could not process file: {att.filename}]",
                    content_block=None,
                )
            )
    return results


def _process_one(att: FileAttachment) -> ProcessedAttachment:
    ct = att.content_type.lower().split(";")[0].strip()

    # --- Images: send as Anthropic vision content block ---
    if ct in _IMAGE_TYPES:
        return ProcessedAttachment(
            filename=att.filename,
            text_content=None,
            content_block={
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": ct,
                    "data": att.data,
                },
            },
        )

    raw = base64.b64decode(att.data)

    # --- PDF ---
    if ct == _PDF_TYPE:
        return _process_pdf(att, raw)

    # --- DOCX ---
    if ct in _DOCX_TYPES:
        return _process_docx(att, raw)

    # --- XLSX ---
    if ct in _XLSX_TYPES:
        return _process_xlsx(att, raw)

    # --- CSV ---
    if ct in _CSV_TYPES or att.filename.lower().endswith(".csv"):
        return _process_csv(att, raw)

    # --- Plain text / markdown ---
    if ct in _TEXT_TYPES or att.filename.lower().endswith((".txt", ".md")):
        return _process_text(att, raw)

    # Fallback: try to treat as text
    try:
        text = raw.decode("utf-8", errors="replace")
        return ProcessedAttachment(
            filename=att.filename,
            text_content=f"[File: {att.filename}]\n{text}",
            content_block=None,
        )
    except Exception:
        return ProcessedAttachment(
            filename=att.filename,
            text_content=f"[Unsupported file type: {att.filename} ({ct})]",
            content_block=None,
        )


# --- Per-type processors ------------------------------------------------------


def _process_pdf(att: FileAttachment, raw: bytes) -> ProcessedAttachment:
    """Try pypdf text extraction; fall back to Claude-native document block."""
    try:
        from pypdf import PdfReader  # type: ignore[import-untyped]

        reader = PdfReader(io.BytesIO(raw))
        pages: list[str] = []
        for i, page in enumerate(reader.pages, 1):
            try:
                text = (page.extract_text() or "").strip()
            except Exception:
                continue
            if text:
                pages.append(f"--- Page {i} ---\n{text}")

        if pages:
            full_text = f"[PDF: {att.filename}]\n" + "\n\n".join(pages)
            return ProcessedAttachment(
                filename=att.filename,
                text_content=full_text,
                content_block=None,
            )
    except ImportError:
        pass
    except Exception:
        logger.debug("pypdf extraction failed for %s, falling back to document block", att.filename)

    # Fallback: send raw PDF as a Claude-native document content block
    return ProcessedAttachment(
        filename=att.filename,
        text_content=None,
        content_block={
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": att.data,
            },
        },
    )


def _process_docx(att: FileAttachment, raw: bytes) -> ProcessedAttachment:
    """Try python-docx text extraction; fall back to a plain note."""
    try:
        from docx import Document  # type: ignore[import-untyped]

        doc = Document(io.BytesIO(raw))
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        if paragraphs:
            text = f"[DOCX: {att.filename}]\n" + "\n\n".join(paragraphs)
            return ProcessedAttachment(
                filename=att.filename,
                text_content=text,
                content_block=None,
            )
    except ImportError:
        pass
    except Exception:
        logger.debug("python-docx extraction failed for %s", att.filename)

    return ProcessedAttachment(
        filename=att.filename,
        text_content=f"[DOCX file uploaded: {att.filename} — text extraction unavailable]",
        content_block=None,
    )


def _process_xlsx(att: FileAttachment, raw: bytes) -> ProcessedAttachment:
    """Try openpyxl to markdown table; fall back to a plain note."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sections: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            table = _rows_to_markdown(rows)
            sections.append(f"### Sheet: {sheet_name}\n{table}")
        wb.close()

        if sections:
            text = f"[XLSX: {att.filename}]\n" + "\n\n".join(sections)
            return ProcessedAttachment(
                filename=att.filename,
                text_content=text,
                content_block=None,
            )
    except ImportError:
        pass
    except Exception:
        logger.debug("openpyxl extraction failed for %s", att.filename)

    return ProcessedAttachment(
        filename=att.filename,
        text_content=f"[XLSX file uploaded: {att.filename} — spreadsheet parsing unavailable]",
        content_block=None,
    )


def _process_csv(att: FileAttachment, raw: bytes) -> ProcessedAttachment:
    """Parse CSV to markdown table using stdlib csv."""
    try:
        text_data = raw.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text_data))
        rows = [tuple(row) for row in reader]
        if rows:
            table = _rows_to_markdown(rows)
            return ProcessedAttachment(
                filename=att.filename,
                text_content=f"[CSV: {att.filename}]\n{table}",
                content_block=None,
            )
    except Exception:
        logger.debug("CSV parsing failed for %s", att.filename)

    return ProcessedAttachment(
        filename=att.filename,
        text_content=f"[CSV file uploaded: {att.filename} — could not parse]",
        content_block=None,
    )


def _process_text(att: FileAttachment, raw: bytes) -> ProcessedAttachment:
    """Direct text inclusion."""
    text = raw.decode("utf-8", errors="replace")
    return ProcessedAttachment(
        filename=att.filename,
        text_content=f"[File: {att.filename}]\n{text}",
        content_block=None,
    )


# --- Helpers ------------------------------------------------------------------


def _rows_to_markdown(rows: list[tuple]) -> str:
    """Convert a list of row tuples into a markdown table string."""
    if not rows:
        return ""

    # Stringify all cells
    str_rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows]

    # Use first row as header
    header = str_rows[0]
    separator = ["---"] * len(header)
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in str_rows[1:]:
        # Pad or trim to header length
        padded = row + [""] * (len(header) - len(row)) if len(row) < len(header) else row[: len(header)]
        lines.append("| " + " | ".join(padded) + " |")
    return "\n".join(lines)
