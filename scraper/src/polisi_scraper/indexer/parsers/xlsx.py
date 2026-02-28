"""XLSX parser preserving sheet and row context."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from polisi_scraper.indexer.parsers.base import DocumentParser, ParsedBlock, ParsedDocument


class XlsxParser(DocumentParser):
    file_type = "xlsx"

    def parse_bytes(
        self,
        payload: bytes,
        *,
        metadata: dict[str, object] | None = None,
    ) -> ParsedDocument:
        workbook = load_workbook(BytesIO(payload), data_only=True)
        blocks: list[ParsedBlock] = []

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_stringify(value) for value in rows[0]]
            for row_number, row in enumerate(rows[1:], start=2):
                values = [_stringify(value) for value in row]
                if not any(values):
                    continue
                row_pairs = []
                for header, value in zip(headers, values):
                    if value:
                        label = header or f"Column {len(row_pairs) + 1}"
                        row_pairs.append(f"{label}: {value}")
                text = " | ".join(row_pairs) if row_pairs else " | ".join(values)
                row_label = values[0] or headers[0] or f"Row {row_number}"
                blocks.append(
                    ParsedBlock(
                        text=text,
                        block_type="row",
                        sheet_name=sheet.title,
                        row_number=row_number,
                        row_label=row_label,
                    )
                )

        return ParsedDocument(
            file_type=self.file_type,
            title=(metadata or {}).get("title") if metadata else None,
            blocks=blocks,
            metadata=dict(metadata or {}),
        )


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
